import re

import openpyxl
import pandas as pd
from functions.mor_levi_compare_functions import change_prices_mor_levi, delete_irrelevant_categorys, finalize_sale_price, change_inventory, change_item_status, extract_products_to_delete, \
    extract_products_to_new, create_upload_to_update_file, clean_up_website_file
from functions.prices_growth_qa import check_min_growths


def create_website_mor_levi_only_file(website_file_path):
    df = pd.read_csv(website_file_path, low_memory=False)
    mor_levi = df.copy()
    mor_levi = clean_up_website_file(mor_levi)
    mor_levi.to_csv("autosync-suppliers\\files\\aio_website_mor_levi_only.csv", encoding='utf-8-sig', index=False)


def create_updated_file(supplier_file_path):
    df = pd.read_csv(supplier_file_path)
    mor_levi = df.copy()

    mor_levi = change_prices_mor_levi(mor_levi)
    mor_levi = delete_irrelevant_categorys(mor_levi)
    mor_levi = finalize_sale_price(mor_levi)
    mor_levi = change_inventory(mor_levi)
    mor_levi = change_item_status(mor_levi)
    mor_levi = check_min_growths(mor_levi)
    mor_levi.to_csv("Files\\mor_levi_final\\mor_levi_updated_prices.csv", encoding='utf-8-sig', index=False)


def create_deleted_file(supplier_file_path, aio_website_file_path):
    df = pd.read_csv(supplier_file_path)
    df2 = pd.read_csv(aio_website_file_path)
    mor_levi = df.copy()
    all_in_one = df2.copy()
    extract_products_to_delete(mor_levi, all_in_one).to_csv("Files\\mor_levi_final\\mor_levi_delete.csv", encoding='windows-1255', index=False)


def create_new_to_upload_file(supplier_file_path, aio_website_file_path):
    df = pd.read_csv(supplier_file_path)
    df2 = pd.read_csv(aio_website_file_path)
    mor_levi = df.copy()
    all_in_one = df2.copy()
    extract_products_to_new(mor_levi, all_in_one).to_csv("Files\\mor_levi_final\\mor_levi_new.csv", encoding='windows-1255', index=False)
    return mor_levi, all_in_one


def create_update_file_to_upload(supplier_file_path, aio_website_file_path):
    df = pd.read_csv(supplier_file_path)
    df2 = pd.read_csv(aio_website_file_path)
    mor_levi = df.copy()
    all_in_one = df2.copy()
    create_upload_to_update_file(mor_levi, all_in_one).to_csv("Files\\mor_levi_final\\mor_levi_update.csv", encoding='windows-1255', index=False)
    return mor_levi, all_in_one


def create_new_items_for_picture_parse(supplier_file_path):
    df = pd.read_csv(supplier_file_path, usecols=['ItemId', 'ItemURL'], encoding='windows-1255')
    mor_levi = df.copy()
    mor_levi.to_csv("Files\\mor_levi_final\\mor_levi_new_items_for_picture_parse.csv", encoding='windows-1255', index=False)
    return mor_levi


def csv_to_xlsx_with_categories(input_csv, output_xlsx):
    df = pd.read_csv(input_csv, encoding='windows-1255')
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    unique_combos = df[['Category', 'SubCategory']].drop_duplicates()
    for _, row in unique_combos.iterrows():
        category = row['Category']
        subcategory = row['SubCategory']
        print(f"category: {category}, subcategory: {subcategory}")
        sheet_name = clean_sheet_name(f"{category}_{subcategory}")
        base_name = sheet_name
        counter = 1
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{base_name}_{counter}"
            counter += 1
        sheet_data = df[(df['Category'] == category) & (df['SubCategory'] == subcategory)]
        worksheet = workbook.create_sheet(title=sheet_name)
        for r_idx, _row in enumerate(sheet_data.itertuples(index=False), start=1):
            for c_idx, value in enumerate(_row, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
        for c_idx, column in enumerate(sheet_data.columns, start=1):
            worksheet.cell(row=1, column=c_idx, value=column)
    workbook.save(output_xlsx)


def clean_sheet_name(name):
    invalid_chars = r'[\\/*?:\[\]]'
    name = re.sub(invalid_chars, '.', name)
    name = name[:31]
    if not name.strip('.'):
        name = "Sheet"
    return name
